"""Generacion de archivos Excel (.xlsx) con openpyxl."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.exportar.reportes import Columna, Reporte

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TOTALES_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_TOTALES_FONT = Font(bold=True)


def _formato_celda(formato: str) -> str | None:
    if formato == "moneda":
        return '"S/ "#,##0.00'
    if formato == "numero":
        return "#,##0"
    if formato == "porcentaje":
        return "0.00"
    if formato == "fecha":
        return "yyyy-mm-dd"
    return None


def _valor_para_celda(valor: Any, formato: str) -> Any:
    if valor is None:
        return None
    if formato == "fecha" and isinstance(valor, datetime):
        return valor.date()
    if formato in ("moneda", "numero", "porcentaje"):
        try:
            return float(valor)
        except (TypeError, ValueError):
            return None
    return valor


def generar(
    reporte: Reporte,
    datos: list[dict[str, Any]],
    *,
    filtros: dict[str, Any] | None = None,
    usuario_nombre: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = reporte.codigo[:31]  # limite openpyxl

    # Metadata en las 3 primeras filas
    ws.cell(row=1, column=1, value=reporte.titulo).font = Font(size=14, bold=True)
    generado = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    if usuario_nombre:
        generado += f" por {usuario_nombre}"
    ws.cell(row=2, column=1, value=generado)
    if filtros:
        filtro_str = " · ".join(f"{k}={v}" for k, v in filtros.items() if v not in (None, ""))
        if filtro_str:
            ws.cell(row=3, column=1, value=f"Filtros: {filtro_str}")

    # Header
    header_row = 5
    for i, col in enumerate(reporte.columnas, start=1):
        c = ws.cell(row=header_row, column=i, value=col.label)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for r_idx, fila in enumerate(datos, start=header_row + 1):
        for c_idx, col in enumerate(reporte.columnas, start=1):
            valor = _valor_para_celda(fila.get(col.key), col.formato)
            celda = ws.cell(row=r_idx, column=c_idx, value=valor)
            fmt = _formato_celda(col.formato)
            if fmt is not None:
                celda.number_format = fmt

    # Totales
    if reporte.con_totales and reporte.columnas_totalizables and datos:
        total_row = header_row + 1 + len(datos)
        ws.cell(row=total_row, column=1, value="TOTAL").font = _TOTALES_FONT
        for i, col in enumerate(reporte.columnas, start=1):
            celda = ws.cell(row=total_row, column=i)
            celda.fill = _TOTALES_FILL
            celda.font = _TOTALES_FONT
            if col.key in reporte.columnas_totalizables:
                total = sum(
                    float(f.get(col.key) or 0)
                    for f in datos
                    if f.get(col.key) is not None
                )
                celda.value = total
                fmt = _formato_celda(col.formato)
                if fmt is not None:
                    celda.number_format = fmt

    # Ancho de columnas: aproximacion por longitud maxima
    for i, col in enumerate(reporte.columnas, start=1):
        max_len = len(col.label)
        for fila in datos[:100]:  # sampling
            v = fila.get(col.key)
            if v is not None:
                max_len = max(max_len, len(str(v)[:60]))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
